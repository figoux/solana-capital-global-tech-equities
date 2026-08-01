#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Solana Global Tech — atualiza a matriz QUANT de correlação entre buckets.

- Baixa retornos históricos via yfinance (rodar no Mac Mini ou no backend do dashboard).
- Cada bucket é um blend equal-weight de proxies (edite PROXIES à vontade).
- Escreve:
    1) aba "Correl Quant" do Excel (mantém formatação; só substitui valores)
    2) static/corr_matrices.json  (quali + quant) para a página do dashboard

Uso:
    pip install yfinance openpyxl pandas
    python update_correlations.py --xlsx Solana_Matriz_Exposicao_Risco_v2.xlsx \
        --json static/corr_matrices.json --period 2y --interval 1wk
"""
import argparse, json, datetime
import pandas as pd
import yfinance as yf
import openpyxl

BUCKETS = ["AI Compute & Semis", "Memória & Storage", "SemiCap (Equip.)", "Cloud & Data Infra",
           "SaaS / Enterprise", "Cybersecurity", "Digital Ads", "E-commerce / Cons. Internet",
           "Fintech / EM", "Consumer HW & Gaming", "Networking & Óptica", "Índice / Mercado"]

# Proxies por bucket (blend equal-weight dos retornos). Edite livremente.
PROXIES = {
    "AI Compute & Semis":          ["SMH"],
    "Memória & Storage":           ["MU", "STX", "WDC"],
    "SemiCap (Equip.)":            ["ASML", "AMAT", "LRCX", "KLAC"],
    "Cloud & Data Infra":          ["SKYY"],
    "SaaS / Enterprise":           ["IGV"],
    "Cybersecurity":               ["CIBR"],
    "Digital Ads":                 ["META", "GOOGL", "TTD"],
    "E-commerce / Cons. Internet": ["AMZN", "MELI", "SE"],
    "Fintech / EM":                ["FINX"],
    "Consumer HW & Gaming":        ["AAPL"],
    "Networking & Óptica":         ["ANET", "CSCO", "CIEN"],
    "Índice / Mercado":            ["SPY"],
}

# célula superior-esquerda da grade de valores na aba (B5), como no arquivo gerado
XL_FIRST_ROW, XL_FIRST_COL = 5, 2
QUALI_SHEET, QUANT_SHEET = "Correl Quali", "Correl Quant"


def compute_quant(period: str, interval: str):
    tickers = sorted({t for ts in PROXIES.values() for t in ts})
    px = yf.download(tickers, period=period, interval=interval,
                     auto_adjust=True, progress=False)["Close"]
    if isinstance(px, pd.Series):
        px = px.to_frame()
    rets = px.pct_change().dropna(how="all")
    bucket_rets = pd.DataFrame({
        b: rets[[t for t in PROXIES[b] if t in rets.columns]].mean(axis=1)
        for b in BUCKETS
    }).dropna()
    corr = bucket_rets.corr().reindex(index=BUCKETS, columns=BUCKETS)
    ticker_corr = rets.dropna(how="any").corr()  # matriz ticker x ticker
    n_obs = len(bucket_rets)
    print(f"{n_obs} observações ({interval}, {period}); {len(tickers)} tickers")
    return corr.round(2), ticker_corr.round(2)


def read_quali(xlsx_path: str) -> list:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[QUALI_SHEET]
    n = len(BUCKETS)
    return [[ws.cell(XL_FIRST_ROW + i, XL_FIRST_COL + j).value for j in range(n)] for i in range(n)]


def write_quant_to_xlsx(xlsx_path: str, corr: pd.DataFrame, ticker_corr: pd.DataFrame):
    wb = openpyxl.load_workbook(xlsx_path)  # nunca usar data_only=True (destruiria fórmulas)
    ws = wb[QUANT_SHEET]
    n = len(BUCKETS)
    for i in range(n):
        for j in range(n):
            ws.cell(XL_FIRST_ROW + i, XL_FIRST_COL + j).value = float(corr.iloc[i, j])
    ws["A3"] = f"Atualizada em {datetime.date.today().isoformat()} via update_correlations.py"

    # aba adicional: correlação ticker x ticker
    tname = "Correl Tickers"
    if tname in wb.sheetnames:
        del wb[tname]
    wt = wb.create_sheet(tname)
    wt["A1"] = f"Correlação ticker × ticker — {datetime.date.today().isoformat()}"
    tk = list(ticker_corr.columns)
    for j, t in enumerate(tk):
        wt.cell(2, 2 + j).value = t
    for i, t in enumerate(tk):
        wt.cell(3 + i, 1).value = t
        for j in range(len(tk)):
            c = wt.cell(3 + i, 2 + j)
            c.value = float(ticker_corr.iloc[i, j])
            c.number_format = "0.00"
    wb.save(xlsx_path)
    print(f"abas '{QUANT_SHEET}' e '{tname}' atualizadas em {xlsx_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="Solana_Matriz_Exposicao_Risco_v2.xlsx")
    ap.add_argument("--json", default="frontend/assets/corr_matrices.json")
    ap.add_argument("--period", default="2y")
    ap.add_argument("--interval", default="1wk")
    a = ap.parse_args()

    quant, ticker_corr = compute_quant(a.period, a.interval)
    write_quant_to_xlsx(a.xlsx, quant, ticker_corr)

    tk = list(ticker_corr.columns)
    payload = {
        "updated": datetime.date.today().isoformat(),
        "period": a.period, "interval": a.interval,
        "buckets": BUCKETS,
        "proxies": PROXIES,
        "quali": read_quali(a.xlsx),
        "quant": [[float(quant.iloc[i, j]) for j in range(len(BUCKETS))] for i in range(len(BUCKETS))],
        "tickers": tk,
        "quant_tickers": [[float(ticker_corr.iloc[i, j]) for j in range(len(tk))] for i in range(len(tk))],
    }
    with open(a.json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print(f"JSON salvo em {a.json}")


if __name__ == "__main__":
    main()
